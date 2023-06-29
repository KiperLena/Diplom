import os
import sqlite3
import openpyxl


def SPM():
    prj_dir = os.path.abspath(os.path.curdir)
    a = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    base_name = 'db.sqlite3'
    connect = sqlite3.connect(prj_dir + '/' + base_name)
    cursor = connect.cursor()
    cursor.execute('CREATE TABLE IF NOT EXISTS centre_area (title text, seria text, num int, type text, otvod text, start int, stop int, region text)')

    file_to_read = openpyxl.load_workbook('licen.xlsx', data_only=True)
    sheet = file_to_read['Sheet1']

    for row in range(2, sheet.max_row + 1):
        # Объявление списка
        data = []
        # Цикл по столбцам от 1 до 4 ( 5 не включая)
        for col in range(1, 10):
            # value содержит значение ячейки с координатами row col
            value = sheet.cell(row, col).value
            # Список который мы потом будем добавлять
            data.append(value)

        # Вставка данных в поля таблицы
        cursor.execute("INSERT INTO centre_area VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?);", (data[0], data[1], data[2], data[3], data[4], data[5], data[6], data[7], data[8]))
    # сохраняем изменения
    connect.commit()
    # закрытие соединения
    connect.close()


def clear_base():
    '''Очистка базы sqlite'''

    # Получаем текущую папку проекта
    prj_dir = os.path.abspath(os.path.curdir)

    # Имя базы
    base_name = 'db.sqlite3'

    connect = sqlite3.connect(prj_dir + '/' + base_name)
    cursor = connect.cursor()

    # Запись в базу, сохранение и закрытие соединения
    cursor.execute("DELETE FROM centre_area")
    connect.commit()
    connect.close()


# Запуск функции
SPM()



